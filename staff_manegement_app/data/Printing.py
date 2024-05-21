from openpyxl import Workbook,load_workbook
import os

class ExcelEditor:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = None

    def __enter__(self):
        if os.path.exists(self.filename):
            self.workbook = load_workbook(self.filename)
        else:
            print(f"ファイル {self.filename} が存在しません。")
        return self.workbook

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.workbook:
            self.workbook.save(self.filename)
            self.workbook.close()

class Notice_Printing:
    def __init__(self,data,set):
        self.set = set
        self.data = data
        from data.SQL_center import Staff_printing_detail
        staff = Staff_printing_detail(data["id"])
        detail = staff.get_data()
        self.excel_printing(detail)

    def excel_printing(self,data):
        from data.SQL_center import Rank_details
        from GUI.load_config import Load_Working_config
        import json
        workbook = Workbook()
        cell_set = Load_Working_config()

        birthday_data = self.era_change(json.loads(data["スタッフ詳細"])['生年月日'])
        
        joining_data = self.era_change(data["入社日"])
        input_joining_data = f'{joining_data[0]}{joining_data[1]}年{joining_data[2]}月{joining_data[3]}日'

        start_date = self.update_date(json.loads(data["更新"])['更新日'],data["期間の定め"])
        
        start_date = self.J_change(start_date)
        
        update_data = self.J_change(json.loads(data["更新"])['更新日'])
        
        progress_time = self.work_time(json.loads(data["勤務時間"])['出勤'],json.loads(data["勤務時間"])['退勤'],json.loads(data["勤務時間"])['休憩'])
        
        over_time_detail = self.over_time_detail(json.loads(data['残業'])['有無'],json.loads(data['残業'])['開始'],json.loads(data['残業'])['終了'])
        rank_detail = Rank_details(data["等級"])
        rank_data = rank_detail.result_get()
        salary_data = self.emp_type(data['雇用形態'],rank_data)
        
        data_dict = list(json.loads(data['主な交通費']).keys())[1]
        caru_set = data_dict[:-1]
        
        xlsx_name = '労働条件通知書(原本).xlsx'
        
        config_path = f'config/{xlsx_name}'
        with ExcelEditor(config_path) as workbook:
            sheet = workbook.active
            
            cell_insert_data = {
                cell_set['emp_type'][0]: f'{salary_data[0]}',
                cell_set["kana"][0]: f'{json.loads(data["氏"])["カナ"]} {json.loads(data["名"])["カナ"]}',  # カナ
                cell_set["name"][0]: f'{json.loads(data["氏"])["氏"]} {json.loads(data["名"])["名"]}',
                cell_set['birthday'][0]: f'{birthday_data[0]}',
                cell_set['birthday'][1]: f'{birthday_data[1]}',
                cell_set['birthday'][2]: f'{birthday_data[2]}',
                cell_set['birthday'][3]: f'{birthday_data[3]}',
                cell_set['joining_day'][0]: f'{input_joining_data}',
                cell_set['post_num'][0]: f'{json.loads(data["スタッフ詳細"])["郵便番号"]}',
                cell_set['address'][0]: f'{json.loads(data["スタッフ詳細"])["住所"]}',
                cell_set['tell'][0]: f'{json.loads(data["スタッフ詳細"])["固定電話"]}',
                cell_set['phone'][0]: f'{json.loads(data["スタッフ詳細"])["携帯電話"]}',
                cell_set['Establishment_of_period'][0]: f'{data["期間の定め"]}',
                cell_set['Est'][0]: f'{start_date}',
                cell_set['Est'][1]: f'{update_data}',
                cell_set['work_place'][0]: f'{data["就業場所"]}',
                cell_set['Job_details'][0]: f'{data["仕事内容"]}',
                cell_set['work_start'][0]: f'{json.loads(data["勤務時間"])["出勤"]}',
                cell_set['work_end'][0]: f'{json.loads(data["勤務時間"])["退勤"]}',
                cell_set['working_time_day'][0]: f'{progress_time[0]}',
                cell_set['working_time_day'][1]: f'{progress_time[1]}',
                cell_set['break_out'][0]: f'{json.loads(data["勤務時間"])["休憩"]}',
                cell_set['week_working_time'][0]: f'{json.loads(data["勤務時間"])["1週間"]}',
                cell_set['holiday'][0]: f'{data["休日"]}',
                cell_set['over_time_period'][0]: f'{over_time_detail[0]}',
                cell_set['over_start'][0]: f'{over_time_detail[1]}',
                cell_set['over_end'][0]: f'{over_time_detail[2]}',
                cell_set['over_time'][0]: f'{over_time_detail[3]}',
                cell_set['over_time'][1]: f'{over_time_detail[4]}',
                cell_set['Salary_form'][0]: f'{salary_data[1]}',
                cell_set['Salary'][0]: salary_data[2],
                cell_set['rank'][0]: f'{data["等級"]}',
                cell_set['Method_of_calculation'][0]: f'{caru_set}',
                cell_set['Method_of_calculation'][1]: f'{json.loads(data["主な交通費"])[f'{caru_set}:']}',
                cell_set['allowance'][0]: f'{salary_data[3]}',
                cell_set['allowance'][1]: f'{salary_data[4]}',
                cell_set['perfect'][0]: f'{salary_data[5]}'
            }
            for position,value in cell_insert_data.items():
                sheet[position] = value
                
        self.print_specific_pages(config_path,1,1)
        
      
        
    
    def era_change(self,date):
        from datetimejp import JDate
        list_date = date.split("-")
        J_date = JDate(int(list_date[0]),int(list_date[1]),int(list_date[2]))
        result = J_date.strftime('%g%e年%m月%d日')
        result_list = [result[:2],result[2:4],result[5:7],result[8:10]]
        return result_list
    
    def update_date(self,date,S_date):
        from datetime import datetime, timedelta
        
        list_date = date.split("-")
        
        li_year = int(list_date[0]) - 1
        
        subtraction = 0
        
        if (li_year % 4 == 0) and ((li_year % 100 != 0) or (li_year % 400 == 0)):
            #print("うるう年")
            subtraction = 364
        else:
            #print("平年")
            subtraction = 363
        data = datetime(year=int(list_date[0]), month=int(list_date[1]), day=int(list_date[2]))
        
        start_date = data - timedelta(days=subtraction)
        
        if S_date == "有":
            pass
        else:
            start_date = ""
        
        #print(start_date)
        
        return start_date

    def J_change(self, date):
        from datetimejp import JDate
        from datetime import datetime
        if isinstance(date, str):
            date = datetime.strptime(date, '%Y-%m-%d')
            
        # JDateに変換
        J_date = JDate(date.year, date.month, date.day)
        
        # 元号形式の日付文字列を取得
        result = J_date.strftime('%g%e年%m月%d日')
        
        return result
    
    
    def work_time(self,start,end,break_time):
        from datetime import datetime, timedelta
        time1 = datetime.strptime(start, '%H:%M')
        time2 = datetime.strptime(end, '%H:%M')
        
        time_difference = time2 - time1
        hours, minutes, seconds = str(time_difference).split(":")
        time3 = 0
        
        if break_time == "1時間":
            time3 = datetime.strptime("01:00", '%H:%M')
        elif break_time == "30分":
            time3 = datetime.strptime("00:30", '%H:%M')
        time_difference = datetime.strptime(f'0{hours}:{minutes}', '%H:%M')
        #print(time_difference)
        return_time = time_difference - time3
        return_time_str = str(return_time)
        date_list = str(return_time_str).split(":")
        return date_list
        
    
    def over_time_detail(self,or_date,start,end):
        data = [or_date,start,end]
        if or_date == '有':
            data[0] = "※原則1時間15分以内の残業があります"
            from datetime import datetime, timedelta
            time1 = datetime.strptime(start, '%H:%M')
            time2 = datetime.strptime(end, '%H:%M')
            time_difference = time2 - time1
            hours, minutes, seconds = str(time_difference).split(":")
            data.append(hours)
            data.append(minutes)
            
            
        else:
            
            data[0] = "※原則残業はありません"
            hours = ""
            minutes = ""
            data.append(hours)
            data.append(minutes)
            
        return data
        
    
    def emp_type(self,type,rank_data):

        #type = "正社員"
        if type == "正社員":
            salary_data = ['【正社員】','月給【基本給】',rank_data[4],rank_data[5],rank_data[6],rank_data[7]]

        else:
            salary_data = ['【パート】','時給',rank_data[9],0,0,0]
        
        return salary_data

            

    def print_specific_pages(self,excel_file_path, from_page, to_page):
        
        from data.SQL_center import setting_select
        import os
        import win32com.client as win32
        abs_path = os.path.abspath(excel_file_path)
        
        printer = setting_select(1)
        printer_name = f'{printer.get_data()[0][0]}'
        
        excel = win32.Dispatch("Excel.Application")
        
        excel.Visible = False
        
        excel.DisplayAlerts = False
        
        if self.set == 'PDF':
            printer_name = 'Microsoft Print to PDF'
        else:
            pass
    
        
        try:
            workbook = excel.Workbooks.Open(abs_path)
            ws = workbook.ActiveSheet
            ws.PageSetup.PrintArea = None  # もし特定の印刷エリアが設定されていればクリア
            ws.PrintOut(From=from_page, To=to_page, Copies=1, Preview=False, ActivePrinter=printer_name)#ここで指定することでエラーが出ない

        except Exception as e:
            print(f"An error occurred: {e}")

        finally:
            
            workbook.Close(SaveChanges=False)
            
            excel.Quit()